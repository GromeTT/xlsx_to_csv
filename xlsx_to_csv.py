# coding: utf-8
#!/usr/bin/env python
import openpyxl
import argparse
import os


def convert_xlsx_to_csv_d(args):
    src_file_path = args.s
    dest_file_path = args.d
    convert_xlsx_to_csv(src_file_path, dest_file_path)


def convert_xlsx_to_csv(src_file_path, dest_file_path):
    if not os.path.isfile(src_file_path):
        print('The file %s does not exist.' % src_file_path)
        return

    work_book = ''
    try:
        work_book = openpyxl.load_workbook(src_file_path, read_only=True)
    except openpyxl.utils.exceptions.InvalidFileException:
        print('The file %s is no regualar spreadsheet.' % src_file_path)
        return

    work_sheet_names = work_book.get_sheet_names()
    if len(work_sheet_names) < 1:
        print('The workbook %s contains no worksheets.' % src_file_path)
        return

    sheet = work_book.get_sheet_by_name(work_sheet_names[0])

    # print('Max column: %s' % sheet.max_column)
    # print('Max row: %s' % sheet.max_row)
    max_row = sheet.max_row
    max_col = sheet.max_column

    dest_file = open(dest_file_path, mode='w')
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            val = sheet.cell(row=r, column=c).value
            if val is None:
                val = ''
            val = correct_string(val)
            dest_file.write(val)
            if c < max_col:
                dest_file.write(',')
        if r < max_row:
            dest_file.write('\n')
    dest_file.close()


def correct_string(val):
    val = val.encode('utf-8')
    if '\n' in val or ',' in val:
        val = '\"' + val + '\"'
    return val


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('s')
    parser.add_argument('d')
    parser.set_defaults(func=convert_xlsx_to_csv_d)

    args = parser.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()

